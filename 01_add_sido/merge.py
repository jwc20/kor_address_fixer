import openpyxl


def merge_all_sheets_to_one(filename):
    # Load the workbook
    wb = openpyxl.load_workbook(filename)

    # Create a new sheet for merged data
    merged_sheet = wb.create_sheet(title="Merged")

    # Iterate over all sheets in the workbook
    for sheet in wb:
        # Skip the newly created 'Merged' sheet during the loop
        if sheet.title == "Merged":
            continue

        # Copy data from each sheet to the 'Merged' sheet
        for row in sheet.iter_rows(values_only=True):
            # Append each row to the merged sheet
            merged_sheet.append(row)

    # Save the modified workbook
    wb.save(filename)
    print("Sheets have been successfully merged into one.")


# Example usage
merge_all_sheets_to_one("kor_addresses_added.xlsx")
